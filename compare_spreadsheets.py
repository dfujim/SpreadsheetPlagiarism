# Object for comparing Microsoft Excel Spreadsheets with extension .xlsx
# Derek Fujimoto
# May 2018

import openpyxl
import numpy as np

# ========================================================================== #
class comparer(object):
    """
        Object for comparing Microsoft Excel Spreadsheets with extension .xlsx.
        
        Usage: 
        
            Construct object: c = comparer(file1,file2)
            Compare files: c.compare([options],[do_print])
                
                options: comma-seperated list of the following
                    meta: compare metadata
                    exact: compare cell values by coordinate
                do_print: if true, print results to stdout
        
        Saves results of comparison to dictionary results. 
        
        Derek Fujimoto
        May 2018 
    """

    # ====================================================================== #
    def __init__(self,file1,file2):
        """
            file1,file2: filename, with path, of spreadsheets to compare
        """
        
        self.book1 = openpyxl.load_workbook(file1)
        self.book2 = openpyxl.load_workbook(file2)
        
        # results
        self.results = result_dict()
        
    # ====================================================================== #
    def cmpr_exact_values(self,do_print=False):
        """
            Compare cell values for exact match. Not intelligent. 
            
            Output: 
                nsame: number of cells which are identical by coordinate. 
                ntotal: number of cells total which are in a shared range. 
                nexcess: number of non-empty cells which are not in a shared 
                        range. 
                        
            Sets to results: 
                nsame, ntotal, nexcess: as described above
                cell_similarity: nsame/ntotal
        """
        
        # track statistics
        nsame = 0
        ntotal = 0
        nexcess = 0
        
        # get sheet names 
        sheet_names1 = self.book1.sheetnames
        sheet_names2 = self.book2.sheetnames
        
        # iterate over sheets
        for sht1nm,sht2nm in zip(sheet_names1,sheet_names2):
            sht1 = self.book1[sht1nm]
            sht2 = self.book2[sht2nm]
            
            # iterate over cells
            sheet1_cells = [cell.value for row in sht1.rows for cell in row]
            sheet2_cells = [cell.value for row in sht2.rows for cell in row]

            # compare cells for which there is identical content
            size = min(len(sheet1_cells),len(sheet2_cells))
            cmpr = np.equal(sheet1_cells[:size],sheet2_cells[:size])
            ntotal += len(cmpr)
            nsame += np.sum(cmpr)
            
            # add content that is in excess from one sheet
            if len(sheet1_cells) > len(sheet2_cells):
                nexcess += np.sum(np.array(sheet1_cells[size:]) != None)
            else:
                nexcess += np.sum(np.array(sheet2_cells[size:]) != None)
            
        # check for excess sheets
        ndiff_sheets = abs(len(sheet_names1)-len(sheet_names2))
        if ndiff_sheets != 0:            
            
            # get exces sheets
            if len(sheet_names1) > len(sheet_names2):
                sheets = [self.book1[n] for n in sheet_names1[len(sheet_names2):]]
            else:
                sheets = [self.book2[n] for n in sheet_names2[len(sheet_names1):]]
            
            # count number of excess cells in excess sheets    
            for sht in sheets:
                cells = [cell.value for row in sht.rows for cell in row]
                nexcess += np.sum(np.array(cells) != None)
                
        # print results
        if do_print:
            print("Shared range cell content exact match: %d/%d (%.2f" % \
                        (nsame,ntotal,float(nsame)/ntotal*100) +\
                  "%) " + "with %d cells in excess." % nexcess)            
        
        # set to self
        self.results['nsame'] = nsame
        self.results['ntotal'] = ntotal
        self.results['nexcess'] = nexcess
        self.results['cell_similarity'] = np.around(float(nsame)/ntotal,4)
        
        return (nsame,ntotal,nexcess)
        
    # ====================================================================== #
    def cmpr_meta(self,do_print=False):
        """
            Compare meta data
            
            Output: 
                mod: boolean, are file last modified times the same?
                create: boolean, are file creation times the same?
            
            Sets to results: 
                same_modifiy_time, same_create_time: as described above
        """
        
        prop1 = self.book1.properties
        prop2 = self.book2.properties
        
        # compare mod time
        mod = prop1.modified == prop2.modified
        
        # compare create time
        create = prop1.created == prop2.created
        
        # print results
        if do_print:
            print('Sheet modification time is identical: %s' % str(mod))
            print('Sheet creation time is identical:     %s' % str(create))
        
        # set to self
        self.results['same_modify_time'] = mod
        self.results['same_create_time'] = create
        
        return (mod,create)

    # ====================================================================== #
    def compare(self,options='meta,exact',do_print=False):
        """
            Run comparisons on the two files
            
            Options: 
                meta: compare metadata
                exact: compare cell values by coordinate
        """
        
        # get options
        options = options.split(',')
        
        # run options
        if 'meta' in options:
            self.cmpr_meta(do_print=do_print)
        
        if 'exact' in options: 
            self.cmpr_exact_values(do_print=do_print)

# ========================================================================== #
class result_dict(dict):
    """
        Pretty formatting and nice retrieval of dictionary items.
    """

    # ====================================================================== #
    def __getattr__(self, name):
        """Allow element access via dot operator"""
        try:
            return self[name]
        except KeyError:
            raise AttributeError(name)

    __setattr__ = dict.__setitem__
    __delattr__ = dict.__delitem__
    
    # ====================================================================== #
    def __repr__(self):
        """Nice representation of results dictionary"""
        
        # get max length of keys
        keys = list(self.keys())
        if len(keys) == 0: return "No results found."
        
        max_key_len = max(map(len,keys))
        
        # get max length of key items
        items = map(str,[self[key] for key in keys])
        max_item_len = max(map(len,items))
        
        # sort keys
        keys.sort()
        
        # make a table
        s = self.__class__.__name__+': \n'
        for k in keys:
            s += "'" + k.ljust(max_key_len) + "': " + \
                 str(self[k]).ljust(max_item_len) + '\n'
        return s
        
    # ====================================================================== #
    def __dir__(self):
        return list(self.keys())
